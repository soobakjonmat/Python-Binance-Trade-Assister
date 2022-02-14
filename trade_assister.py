import tkinter
import openpyxl
import binance.spot
import binance.futures
from spot import Spot
from margin import Margin
import binance.error
from constants import *

api_key = ""
api_secret = ""

class App():
    def __init__(self) -> None:
        self.root = tkinter.Tk()
        self.root.title("Binance Trading Assister by soobakjonmat")
        self.root.geometry("+1150+200")
        self.root.configure(bg=BG_COLOR, padx=15)

        temp = tkinter.Label(text="Choose the Trading Mode:", font=FONT_LARGE, bg=BG_COLOR, fg="white")
        temp.grid(columnspan=2, pady=10)

        self.spot_button = tkinter.Button(text="Spot", font=FONT_MEDIUM, command=lambda: self.check_record("Spot"))
        self.spot_button.grid(column=0, row=1, pady=10)

        self.margin_button = tkinter.Button(text="Margin", font=FONT_MEDIUM, command=lambda: self.check_record("Margin"))
        self.margin_button.config(state="disabled") # delete this after completing margin.py
        self.margin_button.grid(column=1, row=1, pady=10)
        
    def check_record(self, trading_mode):
        self.trading_mode = trading_mode
        self.status_label = tkinter.Label(text="Checking record file...", font=FONT_MEDIUM, bg=BG_COLOR, fg="white")
        self.status_label.grid(pady=10, columnspan=2)

        try:
            self.wb = openpyxl.load_workbook(RECORD_FILE_NAME)
        except FileNotFoundError:
            self.status_label.config(text="Creating record file...")
            self.create_record_file()
            widgets = self.root.winfo_children()
            for widget in widgets:
                widget.destroy()
            tkinter.Label(text="Enter Your API Key:", font=FONT_MEDIUM, bg=BG_COLOR, fg="white").grid(pady=10)
            self.api_key_entry = tkinter.Entry(width=36, font=FONT_SMALL)
            self.api_key_entry.grid(pady=10)
            tkinter.Label(text="Enter Your API Secret:", font=FONT_MEDIUM, bg=BG_COLOR, fg="white").grid(pady=10)
            self.api_secret_entry = tkinter.Entry(width=36, font=FONT_SMALL)
            self.api_secret_entry.grid(pady=10)
            tkinter.Button(text="Submit",  font=FONT_MEDIUM, command=self.set_api_info).grid(pady=10)
        else:
            sheet = self.wb["API Info"]
            self.api_key = sheet[API_KEY_CELL].value
            self.api_secret = sheet[API_SECRET_CELL].value
            self.create_client()

    def create_record_file(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb["Sheet"])
        spot_sheet = self.wb.create_sheet(title="Spot")
        margin_sheet = self.wb.create_sheet(title="Margin")

        sheets = [spot_sheet, margin_sheet]

        for sheet in sheets:
            sheet["A1"] = SPOT_MARGIN_COL_HEADER[0]
            sheet["B1"] = SPOT_MARGIN_COL_HEADER[1]
            sheet["D1"] = SPOT_MARGIN_COL_HEADER[2]
            sheet["E1"] = SPOT_MARGIN_COL_HEADER[3]

            sheet.column_dimensions["A"].width = SPOT_MARGIN_WIDTHS[0]
            sheet.column_dimensions["B"].width = SPOT_MARGIN_WIDTHS[1]
            sheet.column_dimensions["D"].width = SPOT_MARGIN_WIDTHS[2]
            sheet.column_dimensions["E"].width = SPOT_MARGIN_WIDTHS[3]

            sheet[TOTAL_DEPOSIT_CELL].number_format = SPOT_MARGIN_FORMATS[1]
            sheet[LAST_ACCESS_DATE_CELL].number_format = SPOT_MARGIN_FORMATS[2]

        api_sheet = self.wb.create_sheet(title="API Info")
        api_sheet.column_dimensions["A"].width = API_INFO_WIDTH

        self.wb.save("Record.xlsx")

    def set_api_info(self):
        self.api_key = self.api_key_entry.get()
        self.api_secret = self.api_secret_entry.get()
        sheet = self.wb["API Info"]
        sheet["A1"] = self.api_key
        sheet["A2"] = self.api_secret
        self.wb.save(RECORD_FILE_NAME)
        self.create_client()

    def create_client(self):
        try:
            match self.trading_mode:
                case "Spot":
                    self.spot = Spot(self.api_key, self.api_secret, self.root)
                    self.spot.client.account()
                case "Margin":
                    self.margin = Margin(self.api_key, self.api_secret, self.root)
                    self.margin.client.account()
        except binance.error.ClientError as c_error:
            widgets = self.root.winfo_children()
            for widget in widgets:
                widget.destroy()
            self.show_client_error_message(c_error)
            tkinter.Label(text="Quit and try again", font=FONT_MEDIUM, background=BG_COLOR).grid(pady=10)
        except binance.error.ServerError as server_error:
            self.show_server_error_message(server_error)
        else:
            widgets = self.root.winfo_children()
            for widget in widgets:
                widget.destroy()
            self.create_currency_name_check()

    def create_currency_name_check(self):
        temp = tkinter.Label(text="Enter the name of the cryptocurrency (e.g. BTC):", font=FONT_MEDIUM, bg=BG_COLOR, fg="white")
        temp.grid(pady=(10, 0))

        self.crypto_name_entry = tkinter.Entry(width=10, font=FONT_MEDIUM)
        self.crypto_name_entry.grid(pady=(0, 10))

        temp = tkinter.Label(text="Enter the name of the fiat currency (e.g. USDT):", font=FONT_MEDIUM, bg=BG_COLOR, fg="white")
        temp.grid(pady=(10, 0))

        self.fiat_name_entry = tkinter.Entry(width=10, font=FONT_MEDIUM)
        self.fiat_name_entry.grid(pady=(0, 10))
        self.fiat_name_entry.bind("<Return>", self.test_currency_name)

        self.submit_button = tkinter.Button(text="Submit", font=FONT_MEDIUM, command=self.test_currency_name)
        self.submit_button.grid(pady=10)

        self.status_label = tkinter.Label(text="", font=FONT_SMALL, bg=BG_COLOR, fg="white")
        self.status_label.grid(pady=10)

    def test_currency_name(self):
        self.status_label.config(text="Checking Symbol Name...")
        self.root.update()
        self.submit_button.config(state="disabled")
        self.fiat = self.fiat_name_entry.get().upper()
        self.crypto = self.crypto_name_entry.get().upper()
        self.symbol = self.crypto + self.fiat
        self.fiat_name_entry.delete(0, 'end')
        self.crypto_name_entry.delete(0, 'end')
        
        try:
            match self.trading_mode:
                case "Spot":
                    self.spot.client.avg_price(self.symbol)
                case "Margin":
                    self.margin.client.ticker_price(self.symbol)
        except (binance.error.ClientError) as e:
            self.status_label.config(text=f"{e.error_message} Please check the symbol name and try again.")
            self.submit_button.config(state="normal")
        except (binance.error.ParameterRequiredError) as e:
            self.status_label.config(text="Please enter a symbol name and try again.")
            self.submit_button.config(state="normal")
        else:
            widgets = self.root.winfo_children()
            for widget in widgets:
                widget.destroy()
            match self.trading_mode:
                case "Spot":
                    self.spot.initialize(self.symbol, self.crypto, self.fiat)
                case "Margin":
                    self.margin.initialize(self.symbol, self.crypto, self.fiat)

    def show_client_error_message(self, client_error: binance.error.ClientError):
        tkinter.Label(text=client_error.error_code, font=FONT_MEDIUM, bg=BG_COLOR, fg="white").grid(pady=10)
        tkinter.Label(text=client_error.error_message, font=FONT_MEDIUM, bg=BG_COLOR, fg="white").grid(pady=10)

    def show_server_error_message(self, server_error: binance.error.ServerError):
        tkinter.Label(text=server_error.status_code, font=FONT_MEDIUM, bg=BG_COLOR, fg="white").grid(pady=10)
        tkinter.Label(text=server_error.message, font=FONT_MEDIUM, bg=BG_COLOR, fg="white").grid(pady=10)

if __name__ == "__main__":
    app = App()
    app.root.mainloop()